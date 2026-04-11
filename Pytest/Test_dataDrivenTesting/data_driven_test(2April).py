import pytest
import openpyxl

# creating a workbook
wb=openpyxl.Workbook()
# creating a sheet
sheetName="Sheet1"
# checking if the sheet is present
if sheetName in wb.sheetnames:
    ws=wb[sheetName]
    print("Sheet is present")
# creating a new sheet if sheet is not present
else:
    ws=wb.create_sheet(sheetName)
    print("Sheet is not present")
# creating two columns of username and password
ws['A1']="user"
ws['B1']="password"
# adding the data into the sheet
ws.append(["Anushka","1234"])
ws.append(["max","123433"])
ws.append(["molly","342939"])
ws.append(["sonic","42222"])
# fetching the data from the sheet
print(ws['A1'].value)
print(ws['B1'].value)
print(ws['A2'].value)
print(ws['B2'].value)
print(ws['A3'].value)
print(ws['B3'].value)
print(ws['A4'].value)
print(ws['B4'].value)
print(ws['A5'].value)
print(ws['B5'].value)
# using for loop to fetch the data from the sheet
# for i in range(2,ws.max_row+1):
#     for j in range(1,ws.max_column+1):
#         print(ws.cell(row=i,column=j).value,end=" ")
#     print() #to print in separate lines
for row in ws.iter_rows(min_row=2, values_only=True):
    # built in method that iterates through all the rows in the sheet
    print(row)
# saving the workbook
wb.save("sample.xlsx")