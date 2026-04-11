import pytest
import openpyxl

# creating a workbook
wb=openpyxl.Workbook()
# creating a sheet
sheetName="Sheet1"
if sheetName in wb.sheetnames:
    ws=wb[sheetName]
    print("Sheet is present")
else:
    ws=wb.create_sheet(sheetName)
    print("Sheet is not present")

ws['A1']="user"
ws['B1']="password"
wb.save("sample.xlsx")
ws.append(["Anushka","1234"])